import os
import io
import time
import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import requests
from typing import List, Dict, Optional, Tuple, Set

# ===== Translator設定 =====
TRANSLATOR_KEY = os.getenv("translator-key", "")
TRANSLATOR_REGION = os.getenv("translator-region", "japaneast")
TRANSLATOR_ENDPOINT = os.getenv("translator-endpoint", "https://api.cognitive.microsofttranslator.com")

API_VERSION = "3.0"
CHUNK_SIZE = 80          # 1回のリクエストで送る文数（安全側）
MAX_RETRIES = 6
INITIAL_BACKOFF = 1.0
TIMEOUT_SECS = 30

# ---------------------------------------------------------------------
# Translator API 呼び出し
# ---------------------------------------------------------------------
def _translator_headers() -> Dict[str, str]:
    return {
        "Ocp-Apim-Subscription-Key": TRANSLATOR_KEY,
        "Ocp-Apim-Subscription-Region": TRANSLATOR_REGION,
        "Content-Type": "application/json",
    }

def translate_batch(
    texts: List[str],
    to_langs: List[str],
    from_lang: Optional[str] = None,
    text_type: str = "plain",
    category: Optional[str] = None,
) -> Dict[str, List[str]]:
    """
    texts を to_langs それぞれに翻訳して {lang: [translated_texts]} を返す。
    空文字列はそのまま空で返す。
    """
    cleaned = [("" if t is None else str(t)) for t in texts]
    n = len(cleaned)
    results: Dict[str, List[str]] = {lang: [""] * n for lang in to_langs}
    session = requests.Session()

    for start in range(0, n, CHUNK_SIZE):
        end = min(start + CHUNK_SIZE, n)
        batch = cleaned[start:end]
        if all(t == "" for t in batch):
            continue

        body = [{"Text": t} for t in batch]
        params = {"api-version": API_VERSION, "textType": text_type}
        if from_lang and from_lang.lower() != "auto":
            params["from"] = from_lang
        if category:
            params["category"] = category
        to_params = [("to", lang) for lang in to_langs]

        backoff = INITIAL_BACKOFF
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                resp = session.post(
                    f"{TRANSLATOR_ENDPOINT}/translate",
                    params=list(params.items()) + to_params,
                    headers=_translator_headers(),
                    json=body,
                    timeout=TIMEOUT_SECS,
                )
                if resp.status_code == 200:
                    payload = resp.json()
                    for i, item in enumerate(payload):
                        for tr in item.get("translations", []):
                            lang = tr.get("to")
                            text = tr.get("text", "")
                            if lang in results:
                                results[lang][start + i] = text
                    break
                else:
                    if resp.status_code in (429, 500, 502, 503, 504) and attempt < MAX_RETRIES:
                        time.sleep(backoff)
                        backoff = min(backoff * 2, 32)
                        continue
                    raise RuntimeError(f"Translator API error: {resp.status_code} {resp.text}")
            except requests.RequestException as e:
                if attempt < MAX_RETRIES:
                    time.sleep(backoff)
                    backoff = min(backoff * 2, 32)
                    continue
                raise RuntimeError(f"Network error: {e}")
    return results

# ---------------------------------------------------------------------
# シート全体の文字列セルを抽出（座標と値のリスト）
# マージセルは左上セルのみ対象にし、それ以外はスキップ
# ---------------------------------------------------------------------
def collect_string_cells(ws: Worksheet) -> Tuple[List[str], List[Tuple[int, int]]]:
    """
    Worksheet から翻訳対象の文字列セルを収集して
    texts: List[str], coords: List[(row, col)] を返す。
    """
    max_r, max_c = ws.max_row, ws.max_column

    # マージセルの非左上座標をスキップ対象にする
    skip_coords: Set[Tuple[int, int]] = set()
    top_left_coords: Set[Tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        top_left_coords.add((min_row, min_col))
        # 非左上の座標をスキップセットへ
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if not (r == min_row and c == min_col):
                    skip_coords.add((r, c))

    texts: List[str] = []
    coords: List[Tuple[int, int]] = []
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            if (r, c) in skip_coords:
                continue
            val = ws.cell(row=r, column=c).value

            # 文字列のみ対象。数値/日付/None 等は対象外
            if isinstance(val, str):
                texts.append(val)
                coords.append((r, c))
            else:
                # 左上セル以外のマージセルは既にスキップ。
                # 非文字列は翻訳不要なので無視
                continue
    return texts, coords

# ---------------------------------------------------------------------
# ブック全体処理：全シートをループし、言語ごとに翻訳シートを追加
# ---------------------------------------------------------------------
def translate_all_sheets_to_new_tabs(
    xlsx_bytes: bytes,
    to_langs: List[str],
    from_lang: Optional[str] = None,
    text_type: str = "plain",
    category: Optional[str] = None,
) -> io.BytesIO:
    """
    アップロードされたブックの全シートを対象に翻訳を行い、
    「元シート名_言語」という名前の新タブを追加する。
    """
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)

    # 処理中にシート構成が変わるのを防ぐため、最初にシート名のリストを取得
    original_sheet_names = wb.sheetnames

    for sheet_name in original_sheet_names:
        src_ws: Worksheet = wb[sheet_name]

        # 1. 翻訳対象の文字列セル一覧を抽出
        texts, coords = collect_string_cells(src_ws)
        if not texts:
            continue  # 文字列がないシートはスキップ

        # 2. バッチ翻訳
        translations = translate_batch(
            texts=texts,
            to_langs=to_langs,
            from_lang=from_lang,
            text_type=text_type,
            category=category,
        )

        # 3. 言語ごとにシート複製し、翻訳文字列を書き戻す
        for lang in to_langs:
            new_ws = wb.copy_worksheet(src_ws)
            # シート名が重複しないよう、31文字制限を考慮しつつ命名
            new_title = f"{sheet_name}_{lang}"[:31]
            new_ws.title = new_title

            lang_texts = translations[lang]
            for i, (r, c) in enumerate(coords):
                new_ws.cell(row=r, column=c, value=lang_texts[i])

    # 出力
    out_buf = io.BytesIO()
    wb.save(out_buf)
    wb.close()
    out_buf.seek(0)
    return out_buf

# --- 言語オプションの定義 ---
LANG_OPTIONS = {
    "日本語": "ja",
    "中国語（繁体字）": "zh-Tw",
    "英語": "en",
    "韓国語": "ko"
}

# ---------------------------------------------------------------------
# Streamlit UI 修正版
# ---------------------------------------------------------------------
st.set_page_config(page_title="Excel Translator (All Sheets)", page_icon="🌐", layout="centered")
st.title("Excel翻訳（Azure Translator）")
st.caption("ファイル内のすべてのシートを対象に翻訳版を作成します")

# 環境変数チェック
missing = []
if not TRANSLATOR_KEY: missing.append("TRANSLATOR_KEY")
if not TRANSLATOR_REGION: missing.append("TRANSLATOR_REGION")
if missing:
    st.warning(f"環境変数が未設定です：{', '.join(missing)}")

uploaded = st.file_uploader("Excelファイル（.xlsx）をアップロード", type=["xlsx"])

# 言語選択UI
col1, col2 = st.columns(2)
with col1:
    from_lang_label = st.selectbox(
        "翻訳元言語",
        ["自動検出"] + list(LANG_OPTIONS.keys()),
        index=2  # デフォルト：中国語（繁体字）
    )
    from_lang_code = LANG_OPTIONS.get(from_lang_label, None)

with col2:
    to_lang_labels = st.multiselect(
        "翻訳先言語（複数選択可）",
        list(LANG_OPTIONS.keys()),
        default=["日本語"]
    )
    to_langs = [LANG_OPTIONS[label] for label in to_lang_labels]

text_type = st.selectbox("テキスト種別", ["plain", "html"], index=0)
category = st.text_input("Custom Translator カテゴリID（任意）", "")

run_clicked = st.button("全シート翻訳開始", key="translate_button")

if run_clicked:
    if not uploaded:
        st.error("Excelファイルをアップロードしてください。")
    elif not to_langs:
        st.error("翻訳先言語を選択してください。")
    else:
        try:
            with st.spinner("ファイル全体の翻訳を実行中…（シート数により時間がかかる場合があります）"):
                # ファイルポインタを先頭に戻す
                uploaded.seek(0)
                file_data = uploaded.read()

                out_buf = translate_all_sheets_to_new_tabs(
                    xlsx_bytes=file_data,
                    to_langs=to_langs,
                    from_lang=from_lang_code,
                    text_type=text_type,
                    category=(category or None),
                )
            st.success("全シートの翻訳が完了しました！")
            st.download_button(
                label="結果をダウンロード（xlsx）",
                data=out_buf,
                file_name="all_sheets_translated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_translated"
            )
        except Exception as e:
            st.error(f"エラーが発生しました: {e}")
